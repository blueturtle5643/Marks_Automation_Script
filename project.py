import argparse
import os
import time

import re
import csv
import pandas as pd
import pathlib

from itertools import count, groupby

import argparse

import pymongo
from datetime import datetime, timedelta
import openpyxl
from openpyxl.drawing.image import Image
import subprocess

parser = argparse.ArgumentParser()

parser.add_argument("--files", dest = "workFiles", help = "The file needed for input")

parser.add_argument("--xytech", nargs ='+', dest = "xytech_files", help = "The xytech files needed for input")

parser.add_argument("--baselight", nargs ='+', dest = "baselight_files", help = "The baselight files needed for input")

parser.add_argument("--adflame", nargs ='+', dest = "adflame_files", help = "The autodesk flame files needed for input")

parser.add_argument("--verbose", action = "store_true", help = "Toggle switch")

parser.add_argument("--process", dest = "video_file", help = "The video files needed for input")

parser.add_argument("--xls", action = "store_true", help = "Toggle switch")

args = parser.parse_args()


##############################  Purpose: Check if files of each type exist, exit program if they do not

def open_file_check():

    if args.xytech_files is None:

        print("\nMissing Xytech Files.")

        exit(2)

    if args.baselight_files is None and args.adflame_files is None:

        print("\nMissing Baselight or Autodesk Flame Files.")

        exit(2)




##############################  Purpose: Use Xytech to collect useful information into lists, build final.csv

def xytech_process_file(file_structure_array):

    file_length = len(file_structure_array)

    i = 0

    j = 0

    before_semicolon_list = [array_element[:array_element.index(":")] for array_element in file_structure_array if ":" in array_element]    #Create a list for every substring BEFORE ":" in Xytech.txt

    before_semicolon_list.remove('Location')    #HARDCODE   ASK PROFESSOR

    after_semicolon_list = [array_element[array_element.index(":") + 1:] for array_element in file_structure_array if ":" in array_element]     #Create a list for every substring AFTER ":" in Xytech.txt WARNING DOES NOT RECOGNIZE \n

    #line_4_list = [['Show Location', 'Frames to Fix']]  #HARCODE    ASK PROFESSOR

    after_semicolon_list = [re.sub('\n', '', s) for s in after_semicolon_list]  #Removes the "\n" inserted from semi_colon_list

    after_semicolon_list = [s for s in after_semicolon_list if s]   #Removes every '' in the after_semicolon_list

    for i in range(len(after_semicolon_list)):  #Removes every space in after_semicolon_list

        if after_semicolon_list[i].startswith(" "): 
            
            after_semicolon_list[i] = after_semicolon_list[i][1:].lstrip()

    if args.verbose:

        print("")

        #print(f"\n\nInside [xytech_process_file]:")

        #print(f"\nBefore semicolon list: {before_semicolon_list}")

        #print(f"\nAfter semicolon list: {after_semicolon_list}")
    
    return before_semicolon_list, after_semicolon_list
    



##############################  Purpose: HANDLING THE LOCATION AND NOTES PARAGRAPHS AFTER ":"    

def handle_notes_and_locations(file_structure_array, after_semicolon_list):

    i = 0

    j = 0

    file_length = len(file_structure_array)

    for array_element in file_structure_array:  #every line in xytech is stored as an array element, iterate

        if array_element.find(":") != -1:   #if line finds ":"      REDUDENT WAS FOR TESTING

            #word = array_element.split(":")[0]  #was for inserting as a cell for csv

            if array_element.find(":") != -1 and array_element.find(":") + 1 == array_element.find("\n"):   #if a semi colon is detected, and the character next to FIRST SEMICOLON is \n   #HARDCODE

                while i + 1 < file_length and (file_structure_array[i] != "\n" and file_structure_array[i + 1] != "\n"):  #While not at end of the xytech lines array, and the next 2 characters are not "\n"

                    i += 1
                                #iterate
                    j += 1
                    
                    after_semicolon_list.append(file_structure_array[i])    #add the line to the after_semicolon_array

        if j != 0:

            i -= j  
                    #because I moved "index" of the array to add those extra lines, I need to resync back and reset my iterating tracking variables
            j = 0   
        
        i += 1

        #continue   




##############################  Purpose: Since the final.csv file paths are built off of Xytech.txt, store all file paths into a list  

def store_file_paths(after_semicolon_list):
    
    file_path_list = []

    i = 0

    while i < len(after_semicolon_list):
        
        if after_semicolon_list[i][0] == "/":
            
            file_path_list.append(after_semicolon_list.pop(i))

            i = 0

        i += 1

    file_path_list = [re.sub('\n', '', s) for s in file_path_list]  #erase every \n in the file path list

    return file_path_list

    #NEW EDITION AHHHHHHHHHHHHH




##############################  Purpose: Handling the insertion of the first 4 lines of the final.csv   

def form_csv_first_4_lines(after_semicolon_list, before_semicolon_list):

    #f = open('final.csv', 'w')

    #with open('final.csv', 'w') as f:      DESIGN DECISION TO NOT USE WITH OPEN BECAUSE OF INSERTING ROWS FOR LOOP 

    line_4_list = [['Show Location', 'Frames to Fix']]  #HARCODE    ASK PROFESSOR

    #csv_writer = csv.writer(f, delimiter = ',', quotechar = '"', quoting=csv.QUOTE_MINIMAL)

    csv_writer.writerow(before_semicolon_list)  #write first line into csv

    csv_writer.writerow(after_semicolon_list)   #write second line into csv

    csv_writer.writerows(line_4_list)   #write fourth line into csv




##############################  Purpose: Given Baselight or Autodesk Flame files + processed file path list from Xytech, merge filepaths and create ranges

def handle_merging(array_with_numbers, file_path_list, collection_2, variables_list): #pass contents of baselight folder into here

    user_on_file = variables_list[1]

    date = variables_list[2]

    #f = open('final.csv', 'w')

    #writer = csv.writer(f, delimiter = ',', quotechar = '"', quoting = csv.QUOTE_NONNUMERIC, lineterminator="\n") #csv conditions to play around with

    #Read each line from Baselight file

    for line in array_with_numbers:

        line_parse = line.split(" ")
        
        current_folder = line_parse.pop(0)
        
        #sub_folder = current_folder.replace("/images1/starwars", "")    #EDIT

        sub_folder = current_folder.replace("/images1/Avatar", "")    #EDIT
        
        new_location = ""
        
        #Folder replace check
        
        for xytech_line in file_path_list:  #EDIT
            
            if sub_folder in xytech_line:
            
                new_location = xytech_line.strip()
        
        first = ""
        
        pointer = ""
        
        last = ""
        
        for numeral in line_parse:
        
            #Skip <err> and <null>
        
            if not numeral.strip().isnumeric():
        
                continue
        
            #Assign first number
        
            if first == "":
        
                first = int(numeral)
        
                pointer = first
        
                continue
        
            #Keeping to range if succession
        
            if int(numeral) == (pointer+1):
        
                pointer = int(numeral)
        
                continue
        
            else:
        
                #Range ends or no sucession, output
        
                last = pointer
        
                if first == last:
                    
                    if args.verbose:

                        print ("%s %s" % (new_location, first))

                    csv_writer.writerow([new_location] + [first])

                    insert_into_database_2(user_on_file, date, new_location, first, collection_2)   #1
        
                else:
                    
                    if args.verbose:

                        print ("%s %s-%s" % (new_location, first, last))

                    range_string = "{}-{}".format(first, last)

                    csv_writer.writerow([new_location] + [range_string])  

                    insert_into_database_2(user_on_file, date, new_location, range_string, collection_2)    #1
        
                first = int(numeral)
        
                pointer = first
        
                last=""
        
        #Working with last number each line 
        
        last = pointer
        
        if first != "":
        
            if first == last:
        
                if args.verbose:
                    
                    print ("%s %s" % (new_location, first))

                csv_writer.writerow([new_location] + [first])  

                insert_into_database_2(user_on_file, date, new_location, first, collection_2)   #not work
        
            else:
                
                if args.verbose:

                    print ("%s %s-%s" % (new_location, first, last))

                range_string = "{}-{}".format(first, last)

                csv_writer.writerow([new_location] + [range_string])  

                insert_into_database_2(user_on_file, date, new_location, range_string, collection_2)    #1




##############################  Purpose: Extract every "keyword" in file paths for the file_path_list 

def process_files():

    xytech_file_list = args.xytech_files    #all xytech files are now stored in xytech_file_list

    baselight_file_list = args.baselight_files  #all baselight files are now stored in xytech_file_list

    adflame_file_list = args.adflame_files  #all autodesk flame files are now stored in xytech_file_list

    client, collection_1, collection_2 = initialize_database()

    if args.verbose:

        print(f"\nXytech File List: {xytech_file_list}")

        print(f"\nBaselight File List: {baselight_file_list}")

        print(f"\nAutodesk Flame File List: {adflame_file_list}")


    #for file_path_xytech, file_path_baselight in zip(xytech_file_list, baselight_file_list):  #each file_path represents a file in the file list, iterate through each file and process individually        
        
    
    with open(xytech_file_list[0]) as file:    #Store every line in Xytech.txt into an array

        file_structure_array = file.readlines()

        if args.verbose:

            print("")

            #print(f"\n\n\nInside [open_xytech_files], before forming a csv format.")

            print(f"\n\n\nCurrent Xytech folder: {xytech_file_list[0]} \n\n\nCurrent Xytech folder contents: {file_structure_array}")

    #Extract the first item in xytech file list, there should only be one anyways
        
    #filter the xytech file, strip of annoying elements and get ready to format for csv

    a_s_list, b_s_list = xytech_process_file(file_structure_array)

    handle_notes_and_locations(file_structure_array, a_s_list)

    file_path_list = store_file_paths(a_s_list)

    #HARD CODE AHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHH DEAL WITH Xytech_20230325

    file_path_list = [file_path.replace("starwars", "Avatar") for file_path in file_path_list]

    if args.verbose:

        print("")

        #print(f"\n\n\nInside [open_xytech_files], before forming a csv format.")

        print(f"\n\n\nCurrent Xytech folder: {xytech_file_list[0]} \n\n\nCurrent Xytech folder contents: {file_structure_array}")

        #print(f"\n\n\nCurrent Baselight folder: {baselight_file_list} \n\n\nCurrent Baselight folder contents: {baselight_array}")

        print(f"\n\nXytech file path list: {file_path_list}")

        print(f"\n\n\nAfter semicolon list: {a_s_list} \n\n\nBefore semicolon list: {b_s_list}")

    #   ----------------------------- Xytech processing ends here ----------------------------


    form_csv_first_4_lines(a_s_list, b_s_list)


    # ASSUMPTIONS, 1 XYTECH FILE, AS MANY BASELIGHT AND ADFLAMES, ALL SAME DATES

    # CREATE LOOP, ENDS WHEN ALL 3 FILE LISTS ARE NOT NEEDED/DONE

    # Xytech will match with baselight or autoflame, merge the filepaths on them, and output to csv or database and pop from the list

    # when both baselight and autoflame are finished, pop xytech which will end the loop



    while xytech_file_list or baselight_file_list or adflame_file_list:

        variables_list = []

        string_parse_list = []
    
        if xytech_file_list or (baselight_file_list is None and adflame_file_list is None):

            # Process xytech_file_list
            
            file = xytech_file_list.pop(0)
            
            # Perform operations on file from xytech_file_list
            
            #print("Processing xytech file:", file)



        if baselight_file_list:

            with open(baselight_file_list[0]) as file:    #open first baselight file from baselight list

                baselight_array = file.readlines()  
            
            # Process baselight_file_list

            if args.verbose:

                print("")

                #print(updated_lines)

                print(f"\nBaselight File: {baselight_file_list[0]}\n\n")

                #print(f"\nBaselight File Contents: {baselight_array}")
            
            string_parse_list = baselight_file_list[0]

            variables_list = [part.split('.')[0] if i == 2 else part for i, part in enumerate(string_parse_list.split('_'))]

            #print(variables_list)

            insert_into_database_1(variables_list[0], variables_list[1], variables_list[2], collection_1)

            handle_merging(baselight_array, file_path_list, collection_2, variables_list)
            
            del baselight_file_list[0]  # Should remove the first object in list



        if adflame_file_list:
            
            with open(adflame_file_list[0]) as file:    #open first autodesk flame file from autodesk flame list

                adflame_array = file.readlines()  
            
            #TRIM AD FLAME FILE SO THAT IT CAN BE PROCESSED PROPERLY

            updated_lines = [line.replace('/net/flame-archive ', '') for line in adflame_array]

            if args.verbose:

                print("")

                #print(updated_lines)

                print(f"\nAutodesk Flame File: {adflame_file_list[0]}\n\n")

            # Process adflame_file_list
            
            string_parse_list = adflame_file_list[0]

            variables_list = [part.split('.')[0] if i == 2 else part for i, part in enumerate(string_parse_list.split('_'))]

            insert_into_database_1(variables_list[0], variables_list[1], variables_list[2], collection_1)

            handle_merging(updated_lines, file_path_list, collection_2, variables_list)
            
            del adflame_file_list[0]  # Should remove the first object in list

        if args.verbose:

            print("")

            #print(f"\n\nInside [open_xytech_files], after forming a csv format.")
    
            #print(f"\n\n\nCurrent Xytech folder: {file_path} \n\n\nCurrent Xytech folder contents: {xytech_folders}")

            #print(f"\n\nCurrent Xytech folder: {file_path} \n\nCurrent Xytech folder contents: {file_structure_array}")

            #print(f"\n\nAfter semicolon list: {a_s_list} \n\nBefore semicolon list: {b_s_list}")
    
    client.close()




##############################  Purpose: Create a mango database

def initialize_database():

    client = pymongo.MongoClient("mongodb://localhost:27017/")

    db = client["project_2"]

    collection_1 = db["collection_1"] 
    
    collection_2 = db["collection_2"]

    #project_questions(collection_1, collection_2)

    database_calls(collection_1, collection_2)

    return client, collection_1, collection_2

    # Collection 1: Insert data

    # Close MongoDB connection
    
    #client.close()




##############################  Purpose: Insert into collection 1

def insert_into_database_1(machine, user_on_file, date, db_collection_1):

    db_user = os.getlogin()

    db_name_of_machine = machine

    db_name_of_user_on_file = user_on_file
    
    db_date_of_file = date
    
    db_submitted_date = datetime.now()

    data_1 = {

        "user": db_user,
        "machine": db_name_of_machine,
        "Ñame_of_user_on_file": db_name_of_user_on_file,
        "date_of_file": db_date_of_file,
        "Submitted_date": db_submitted_date
    }

    # Insert data into Collection 1
    
    db_collection_1.insert_one(data_1)




##############################  Purpose: Insert into collection 2

def insert_into_database_2(user_on_file, date, location, frame_ranges, db_collection_2):

    db_name_of_user_on_file = user_on_file
    
    db_date_of_file = date

    db_location = location
    
    db_frame_ranges = frame_ranges
    
    data2 = {}

    data2 = {

        "name_of_user_on_file": db_name_of_user_on_file,
        "date_of_file": db_date_of_file,
        "location": db_location,
        "frame_ranges": db_frame_ranges
    }

    # Insert data into Collection 2
    
    db_collection_2.insert_one(data2)




##############################  Purpose: Project 2 questions

def project_questions(collection_1, collection_2):

    question_1 = collection_2.find({"name_of_user_on_file": "TDanza"})

    print(f"\nQuestion 1: ")

    for i in question_1:

        print(i)
    

    #question_2 = collection_2.find({"date_of_file": "20230323"}, {"location": "Flame"})

    # Print the documents that match the filter criteria

    print(f"\nQuestion 2: ")

    filter_criteria = '20230323'

    # Perform the aggregation

    result = collection_1.aggregate([
        {
            '$lookup': {
                'from': collection_2.name,  # Specify the name of the collection to join with
                'localField': 'date_of_file',  # Specify the field in "collection_1" to use for the join
                'foreignField': 'date',  # Specify the field in "collection_2" to use for the join
                'as': 'matched_docs'  # Specify the name of the field to store the matched documents
            }
        },
        {
            '$match': {
                'machine': 'Flame',
                'matched_docs.date': {'$lt': filter_criteria}  # Specify the filter criteria
            }
        }
    ])

    # Print the documents that match the filter criteria
    
    print("Documents matching filter criteria:")
    
    for matched in result:
        print(matched)



    question_3 = collection_2.find({"name_of_user_on_file": "hpsans13", "date_of_file": "20230326"})

    print(f"\nQuestion 3: ")

    for j in question_3:

        print(j)

    question_4 = collection_1.find({'machine': 'Flame'}, {'Ñame_of_user_on_file': 1})

    print(f"\nQuestion 4: ")

    for i in question_4:

        print(i['Ñame_of_user_on_file'])




##############################  Purpose: Database calls

def database_calls(collection_1, collection_2):

    video_file_name = args.video_file

    #video_file_name = fr"D:\Downloads\467\\twitch_nft_demo.mp4"

    if video_file_name is None:

        print("LALALALLALALALALAL")

    video_timecode = extract_timecode(video_file_name)

    video_timecode_to_seconds = timecode_to_seconds(video_timecode)

    if args.xls:

        wb = openpyxl.Workbook()
        
        ws = wb.active

        ws.cell(row = 1, column = 1, value = "Location")
        
        ws.cell(row = 1, column = 2, value = "Frame Ranges")
        
        ws.cell(row = 1, column = 3, value = "Timecode Ranges")

        ws.cell(row = 1, column = 4, value = "Thumbnails")

        row = 2

        iterator = 1

    for doc in collection_2.find(): 
    
        xls_location = doc["location"]
    
        xls_frame_ranges = doc["frame_ranges"]
    
        frame_range = doc.get("frame_ranges")

        frame_range = str(frame_range)
        
        if frame_range:
                
            if frame_range.isdigit():

                pass

                """ singular_frame = int(frame_range)

                singular_timecode = frame_to_timecode(singular_frame)

                singular_timecode_to_seconds = timecode_to_seconds(singular_timecode)

                if singular_timecode_to_seconds < video_timecode_to_seconds:

                    print(f"\nSingular frame: {singular_frame}, singular timecode: {singular_timecode}") """

                    #record in database
            

            else:
                
                # extract the lowest and highest frame numbers from the range
                
                frame_nums = frame_range.split("-")
                
                lowest_frame = int(frame_nums[0])
                
                highest_frame = int(frame_nums[1])
                
                # convert the frame numbers to timecodes
                
                lowest_timecode = frame_to_timecode(lowest_frame)
                
                highest_timecode = frame_to_timecode(highest_frame)

                lowest_timecode_to_seconds = timecode_to_seconds(lowest_timecode)

                highest_timecode_to_seconds = timecode_to_seconds(highest_timecode)

                median_frame = median(lowest_frame, highest_frame)

                median_timecode = frame_to_timecode(median_frame)


                if lowest_timecode_to_seconds > video_timecode_to_seconds:

                    print(f"\nLowest frame: {lowest_frame}, Lowest timecode: {lowest_timecode}")

                    continue

                    #something database

                
                if highest_timecode_to_seconds > video_timecode_to_seconds:
                    
                    print(f"\nHighest frame: {highest_frame}, Highest timecode: {highest_timecode}")

                    continue

                    #something database

                
                #median_timecode_command = fr"ffmpeg -i input.flv -ss 00:00:14.435 -frames:v 1 out.png"

                thumbnail_name = fr"{iterator}.jpg"
                
                #subprocess.call(['ffmpeg', '-ss', median_timecode, '-i', video_file_name, '-vframes', '1', thumbnail_name])

                file_path = fr"D:\Downloads\467"

                streams = fr"ffmpeg -ss {median_timecode} -i {video_file_name} -vf scale=96:74 -vframes 1 {file_path}\{thumbnail_name}"

                process = subprocess.run(streams, stdout = subprocess.PIPE, stderr = subprocess.STDOUT, shell = True, universal_newlines = True)

                img = Image(fr'{file_path}\{thumbnail_name}')

                if args.xls:

                    xls_time_code = f"{lowest_timecode}-{highest_timecode}"

                    ws.cell(row = row, column = 1, value = xls_location)
                    
                    ws.cell(row = row, column = 2, value = xls_frame_ranges)
                    
                    ws.cell(row = row, column = 3, value = xls_time_code)

                    #ws.cell(row = row, column = 4, value = img)

                    ws.add_image(img, fr"{ws.cell(row = row, column = 4).coordinate}")

                    #ws.cell(row=i, column=4, value=thumbnail_paths[i-2])

                    row += 1

                    iterator += 1
            
                # insert the timecodes into a new "timecodes" field in the document
                
                #collection_2.update_one({"_id": doc["_id"]}, {"$set": {"timecodes": f"{lowest_timecode}-{highest_timecode}"}})

    if args.xls:

        wb.save(r"D:\Downloads\Work\Comp 467\12 - Project 3\\output.xlsx")




##############################  Purpose: Converts frame to Timecode

def frame_to_timecode(frame_num):

    fps = 24

    conversion = frame_num // fps
    
    seconds = conversion % 60

    minutes = (conversion % 3600) // 60

    hours = conversion // 3600

    frames = frame_num % fps
    
    timecode = f"{hours:02d}:{minutes:02d}:{seconds:02d}.{frames:02d}"
    
    return timecode

    #print(f"\nFrame Number: {frame_num} \t Timecode: {timecode}")




##############################  Purpose: Extract timeframe from video

def extract_timecode(video_file_name):

    #test4 = r"ffmpeg -i D:\Downloads\467\\twitch_nft_demo.mp4 -f null -"

    video_file_command = fr"ffmpeg -i {video_file_name} -f null -"

    process = subprocess.run(video_file_command, stdout = subprocess.PIPE, stderr = subprocess.STDOUT, shell = True, universal_newlines = True)

    """ for line in process.stdout.split('\n'):
        print(line) """

    output_lines = process.stdout.splitlines()

    last_line = output_lines[-1]

    match = re.search(r"time=(\d{2}):(\d{2}):(\d{2}.\d{2})", last_line)

    if match:

        time_value = match.group(1) + ":" + match.group(2) + ":" + match.group(3)
        
        #print("Time value:", time_value)
    
    else:
        
        #print("No match found")
        print

    return time_value

    


##############################  Purpose: Converts timecode to seconds

def timecode_to_seconds(timecode_string):
    
    timecode = datetime.strptime(timecode_string, "%H:%M:%S.%f").time()

    total_seconds = timedelta(hours=timecode.hour, minutes=timecode.minute, seconds=timecode.second, microseconds=timecode.microsecond).total_seconds()

    return total_seconds




##############################  Purpose: Find central frame

def median(min_val, max_val):

    values = list(range(min_val, max_val+1))

    sorted_values = sorted(values)

    num_values = len(sorted_values)

    if num_values % 2 == 1:
        
        return sorted_values[num_values//2]
    
    else:
        middle = num_values//2

        return round((sorted_values[middle-1] + sorted_values[middle])/2)




if __name__ == "__main__":  #create main function parallel to c++ main

    f = open('final.csv', 'w')

    csv_writer = csv.writer(f, delimiter = ',', quotechar = '"', quoting = csv.QUOTE_NONNUMERIC, lineterminator="\n")

    initialize_database()

    #open_file_check()

    #process_files()

    f.close()



 